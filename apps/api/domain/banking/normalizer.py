"""
Bank statement normalization engine (Banking B.1, Part C).

TODO(compliance): docs/compliance/05-bank-data-and-the-account-aggregator.md
    This is the ONLY way bank data enters the platform, and it stays that way
    longer than CLAUDE.md's bank-data section assumed. Two verified findings
    (2026-09-04) worth reading before anyone starts an AA feature:

    (1) CLAUDE.md says "Register as an FIU". THAT IS NOT ACHIEVABLE. The RBI
        NBFC-AA Directions 2025 define an FIU as "an entity registered with and
        regulated by any financial sector regulator" — RBI, SEBI, IRDAI, PFRDA
        or the Department of Revenue. There is no FIU licence to apply for and
        no unregulated tier; eligibility is derivative of a registration you
        already hold, and a TSP cannot confer it because a TSP is itself
        unregulated. The Department of Revenue does not open a door: it is in
        that list because it regulates GSTN for the specific purpose of GSTN
        being an FIP. ICAI is not a financial sector regulator either, so a CA
        firm does not qualify. The options are: partner with a regulated FIU,
        acquire a registration, or do not consume via AA.

    (2) Coverage independently confirms upload stays at parity. Co-operative
        banks, RRBs and small finance banks are largely not AA-enabled, and
        even at live banks coverage is patchy BY ACCOUNT TYPE — fixed and
        recurring deposits at only ~40% of banks, joint and non-individual
        accounts worse. A CA's client base is exactly the population AA serves
        worst. Upload is the base case, not a hedge.


Converts CSV / XLSX exports from Indian banks into ONE internal format
(NormalizedTxn: date, description, reference, debit, credit, balance). All
bank-specific column layouts live in `_ADAPTERS` — no bank-specific logic exists
anywhere else (the dispatcher and parsers are format-agnostic).

Money is integer paise via Decimal (never binary float — CLAUDE.md). Dates are
normalised to ISO YYYY-MM-DD.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Optional

_logger = logging.getLogger("caflow.banking.normalizer")


class StatementParseError(ValueError):
    """Raised when a file cannot be parsed into transactions (malformed / unsupported)."""


@dataclass(frozen=True)
class NormalizedTxn:
    transaction_date: str          # ISO YYYY-MM-DD
    description: str
    reference_no: Optional[str]
    debit_paise: int
    credit_paise: int
    balance_paise: int


# ── Bank adapters — the ONLY place that knows bank-specific column layouts ─────
# Each adapter maps a detected format to 0-based column indices. "ref"/"balance"
# may be None when a bank export omits them. An adapter with an "amount" key
# (instead of separate "debit"/"credit" keys) uses a single signed-amount
# column plus a separate Dr/Cr indicator column — see "generic_amount_drcr".
_ADAPTERS: dict[str, dict[str, Optional[int]]] = {
    # Date, Narration, Value Dt, Ref No, Debit, Credit, Balance
    "hdfc":    {"date": 0, "desc": 1, "ref": 3, "debit": 4, "credit": 5, "balance": 6},
    # Txn Date, Value Date, Description, Ref/Cheque No, Debit, Credit, Balance
    "sbi":     {"date": 0, "desc": 2, "ref": 3, "debit": 4, "credit": 5, "balance": 6},
    # Transaction Date, Value Date, Transaction Remarks, Ref No, Debit, Credit, Balance
    "icici":   {"date": 0, "desc": 2, "ref": 3, "debit": 4, "credit": 5, "balance": 6},
    # Tran Date, CHQNO, Narration, Debit, Credit, Balance
    "axis":    {"date": 0, "ref": 1, "desc": 2, "debit": 3, "credit": 4, "balance": 5},
    # Generic: Date, Description, Debit, Credit, Balance
    "generic": {"date": 0, "desc": 1, "ref": None, "debit": 2, "credit": 3, "balance": 4},
    # R2.11: some exports use ONE signed-amount column plus a separate Dr/Cr
    # indicator instead of separate Debit/Credit columns — Date, Description,
    # Amount, Dr/Cr, Balance. Previously unsupported: every row misparsed as a
    # debit (the fixed debit/credit indices pointed at the wrong columns, or
    # at nothing at all).
    "generic_amount_drcr": {
        "date": 0, "desc": 1, "ref": None,
        "debit": None, "credit": None, "amount": 2, "drcr": 3, "balance": 4,
    },
}


def detect_format(headers: list[str]) -> str:
    """Pick a bank adapter from the header row.

    Order matters (F8): a cheque/reference column exists under DIFFERENT names in
    several banks — HDFC "Chq/Ref No", Axis "CHQNO", SBI "Ref/Cheque No" — so the
    generic "chq"/"cheque" heuristic must run LAST. Matching it first routed every
    Axis and SBI statement to the HDFC adapter, and Axis's different column order
    then corrupted debit/credit direction and amounts (a ₹500 debit became a
    ₹10,000 credit). Bank-specific date/remarks columns are the reliable
    discriminators, so they are tested first, most-specific to least.
    """
    h = [str(x).lower().strip() for x in headers]
    blob = " ".join(h)
    # A combined Dr/Cr indicator column header is a signal for the single-
    # amount layout — but ONLY when the file does NOT also have separate,
    # clearly-labelled Debit AND Credit amount columns of its own (adversarial
    # review, R2.11 fix phase): a statement can have separate Debit/Credit
    # columns AND an unrelated "Dr/Cr" column marking the running balance's
    # polarity, not the transaction's direction — routing that layout to the
    # amount+indicator adapter misreads the balance-polarity column as a
    # transaction indicator and fails every row. Requiring the ABSENCE of a
    # separate debit-token cell and a separate credit-token cell keeps this
    # check from firing on that shape, while still catching the true
    # single-amount layout (checked first, same as ICICI's equally
    # unambiguous "transaction remarks" signal below).
    has_drcr_column = any(re.fullmatch(r"dr\s*/\s*cr|dr\s+or\s+cr|cr\s*/\s*dr", x) for x in h)
    has_separate_debit_credit_columns = (
        any(any(t in cell for t in _DEBIT_TOKENS) for cell in h)
        and any(any(t in cell for t in _CREDIT_TOKENS) for cell in h)
    )
    if has_drcr_column and not has_separate_debit_credit_columns:
        return "generic_amount_drcr"
    if "transaction remarks" in blob:      # ICICI ("Transaction Remarks")
        return "icici"
    if "txn date" in blob:                 # SBI ("Txn Date")
        return "sbi"
    if "tran date" in blob:                # Axis ("Tran Date"); note: distinct
        return "axis"                      # from SBI "txn date" and ICICI "transaction date"
    if "narration" in blob or "chq" in blob or "cheque" in blob:  # HDFC (shared cheque/narration signal, last)
        return "hdfc"
    return "generic"


_DEBIT_TOKENS = ("debit", "withdrawal", "withdraw")
_CREDIT_TOKENS = ("credit", "deposit")
_DRCR_HEADER_TOKENS = ("dr", "cr", "type")


def _validate_adapter(fmt: str, headers: list[str]) -> None:
    """Fail loud when the detected adapter does not actually fit the header.

    detect_format uses fuzzy header signals; if a file's real layout differs from
    the chosen adapter — an unsupported bank, or a variant with a shifted column
    order — the fixed index map would silently read debit/credit off the wrong
    columns (the F8 class of corruption). So verify the adapter fits: every mapped
    column exists within the header width, and the debit/credit columns really are
    labelled debit/withdrawal and credit/deposit. Otherwise raise, so a CA sees an
    'unsupported format' error instead of silently wrong numbers.
    """
    a = _ADAPTERS[fmt]
    n = len(headers)
    low = [str(x).lower().strip() for x in headers]
    for key, idx in a.items():
        if idx is not None and idx >= n:
            raise StatementParseError(
                f"Bank statement layout doesn't match the detected '{fmt}' format "
                f"(needs a '{key}' column at position {idx + 1}, but the file has {n} columns). "
                "Supported: HDFC, SBI, ICICI, Axis, a generic "
                "Date/Description/Debit/Credit/Balance CSV, or a generic "
                "Date/Description/Amount/Dr-Cr/Balance CSV."
            )
    if a.get("amount") is not None:
        ai, di = a["amount"], a["drcr"]
        amount_hdr = low[ai] if ai is not None and ai < n else ""
        drcr_hdr = low[di] if di is not None and di < n else ""
        if "amount" not in amount_hdr or not any(t in drcr_hdr for t in _DRCR_HEADER_TOKENS):
            raise StatementParseError(
                "Unsupported bank statement format — could not identify the amount "
                "and Dr/Cr columns. Supported: HDFC, SBI, ICICI, Axis, a generic "
                "Date/Description/Debit/Credit/Balance CSV, or a generic "
                "Date/Description/Amount/Dr-Cr/Balance CSV."
            )
        return
    di, ci = a["debit"], a["credit"]
    debit_hdr = low[di] if di is not None and di < n else ""
    credit_hdr = low[ci] if ci is not None and ci < n else ""
    if not any(t in debit_hdr for t in _DEBIT_TOKENS) or not any(t in credit_hdr for t in _CREDIT_TOKENS):
        raise StatementParseError(
            "Unsupported bank statement format — could not identify the debit and "
            "credit columns. Supported: HDFC, SBI, ICICI, Axis, a generic "
            "Date/Description/Debit/Credit/Balance CSV, or a generic "
            "Date/Description/Amount/Dr-Cr/Balance CSV."
        )


# ── value parsers ─────────────────────────────────────────────────────────────

_MONTHS = {m: f"{i:02d}" for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def _to_iso_date(val) -> Optional[str]:
    """Normalise a date cell to ISO YYYY-MM-DD, or None if unrecognised."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    # D/M/YYYY, DD-MM-YYYY, DD.MM.YYYY (single- or double-digit day/month, any
    # of the three separators a bank export might use) -- widened from a
    # DD-only/2-separator regex that silently dropped otherwise-valid
    # transaction rows whose date happened to use a dot separator or an
    # unpadded single-digit day/month.
    m = re.fullmatch(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", s)
    if m:
        day, mon, year = int(m.group(1)), int(m.group(2)), m.group(3)
        if 1 <= day <= 31 and 1 <= mon <= 12:
            return f"{year}-{mon:02d}-{day:02d}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):                      # already ISO
        return s
    m = re.fullmatch(r"(\d{1,2})[ \-]([A-Za-z]{3})[ \-](\d{4})", s)  # D MMM YYYY
    if m:
        mon_name = _MONTHS.get(m.group(2).lower())
        if mon_name:
            return f"{m.group(3)}-{mon_name}-{int(m.group(1)):02d}"
    return None


_DRCR_SUFFIX_RE = re.compile(r"(dr|cr)$", re.IGNORECASE)


def _to_paise(val) -> int:
    """Parse a money cell to integer paise. Empty / '-' → 0. Never uses float.

    R2.11: a trailing Dr/Cr suffix is common across Indian bank exports in ANY
    case ('Dr', 'DR', 'dr', 'Cr', 'CR', 'cr') — the previous `.rstrip("DrCr")`
    stripped individual CHARACTERS 'D','r','C' (case-sensitive, so it happened
    to handle mixed-case "Dr"/"Cr" but silently zeroed every other case
    variant, since e.g. "150.00DR" has no matching trailing chars to strip and
    Decimal("150.00DR") then raises, caught below, returning 0). Fixed via a
    case-insensitive regex match on the two-letter suffix itself.

    The suffix also carries sign information a bank statement's BALANCE column
    relies on: "Dr" denotes an overdrawn (negative) balance, "Cr" a normal
    (positive) one — previously discarded entirely (both suffixes were just
    stripped, so a "Dr" balance was stored as if it were positive/"Cr"). The
    debit_paise/credit_paise columns wrap this in abs() at the call site, so
    applying the sign here is harmless for them and only matters for balance.

    Adversarial-review fix (R2.11 fix phase): the Dr/Cr suffix is located and
    stripped BEFORE the parens check (not after), so a suffix trailing OUTSIDE
    a parenthesised amount — e.g. "(150.00) Dr" — no longer leaves an orphaned
    ")" that fails to parse. The numeric magnitude is also parsed and made
    absolute BEFORE any sign is applied, rather than negating whatever sign
    Decimal happened to find in the source string — so an amount that already
    carries an explicit leading "-" (e.g. "-150.00 Dr") is not double-negated.
    When a Dr/Cr suffix is present it is authoritative over any other sign
    signal (parens or an embedded "-"), since it is the statement's own
    explicit accounting label; parens/explicit "-" are only used as a
    fallback sign source when no suffix is present.
    """
    if val is None:
        return 0
    if isinstance(val, (int,)):
        return int(val) * 100
    if isinstance(val, float):
        return int((Decimal(str(val)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    s = str(val).strip()
    if s in ("", "-"):
        return 0
    s = re.sub(r"[₹,\s]", "", s)
    m = _DRCR_SUFFIX_RE.search(s)
    drcr = None
    if m:
        drcr = m.group(1).lower()
        s = s[:m.start()]
    if not s:
        return 0
    parens_neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    if not s:
        return 0
    try:
        raw = Decimal(s)
    except InvalidOperation:
        return 0
    explicit_neg = raw < 0
    try:
        paise = int((abs(raw) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return 0
    if drcr == "dr":
        neg = True
    elif drcr == "cr":
        neg = False
    else:
        neg = parens_neg or explicit_neg
    return -paise if neg else paise


def _looks_like_header(cells: list[str]) -> bool:
    low = " ".join(str(c).lower() for c in cells)
    return "date" in low and ("debit" in low or "amount" in low or "withdrawal" in low or "credit" in low)


_INDICATOR_DEBIT_TOKENS = ("d", "dr", "debit", "withdrawal")
_INDICATOR_CREDIT_TOKENS = ("c", "cr", "credit", "deposit")


def _rows_to_txns(rows: list[list], header_idx: int,
                  mapping: Optional[dict] = None) -> list[NormalizedTxn]:
    headers = [str(c) for c in rows[header_idx]]
    if mapping is not None:
        # An explicit mapping DELIBERATELY skips _validate_adapter. That
        # function's job is to catch a mis-DETECTED adapter by checking the
        # bank's column labels really say debit/credit — and the entire reason
        # a mapping exists is that this bank's labels are ones we do not
        # recognise. Requiring them here would refuse every file the mapper is
        # for.
        #
        # What replaces the label check is not nothing, and it is stronger:
        # validate_mapping has already checked the shape, the CA has SEEN the
        # parsed rows in the preview, and balance_agreement() checks the result
        # against the bank's own arithmetic — a swapped debit/credit fails that
        # on the first row, which no label check could catch anyway.
        a = validate_mapping(mapping, len(headers))
    else:
        fmt = detect_format(headers)
        _validate_adapter(fmt, headers)
        a = _ADAPTERS[fmt]
    amount_mode = a.get("amount") is not None
    out: list[NormalizedTxn] = []
    for cells in rows[header_idx + 1:]:
        if not cells or len([c for c in cells if str(c).strip()]) < 3:
            continue

        def col(key: str):
            i = a.get(key)
            if i is None or i >= len(cells):
                return None
            return cells[i]

        iso = _to_iso_date(col("date"))
        desc = (str(col("desc")).strip() if col("desc") is not None else "")
        if not iso or not desc:
            # A row with a real description but an unparseable date is very
            # likely a genuine transaction row, not the "totals/blanks/
            # sub-headers" this skip is meant to filter -- surface it in the
            # server log so a systematic date-format gap (e.g. a bank export
            # using a separator/format _to_iso_date doesn't recognise) is
            # diagnosable instead of silently vanishing transactions.
            if not iso and desc:
                _logger.warning(
                    "bank statement row skipped: description %r present but date %r unrecognised",
                    desc[:80], col("date"),
                )
            continue  # skip non-transaction rows (totals, blanks, sub-headers)
        ref = col("ref")

        if amount_mode:
            # R2.11: single signed-amount + separate Dr/Cr indicator column —
            # classify by the indicator rather than by column position, since
            # there is no separate debit/credit column to read. Trailing
            # punctuation ("Dr.", "Cr.") is stripped (adversarial-review fix)
            # so a bank's own dotted abbreviation isn't silently unrecognised
            # and its rows dropped with no error.
            amount = abs(_to_paise(col("amount")))
            indicator = str(col("drcr") or "").strip().lower().rstrip(".")
            is_debit = indicator in _INDICATOR_DEBIT_TOKENS
            is_credit = indicator in _INDICATOR_CREDIT_TOKENS
            if amount == 0 or (not is_debit and not is_credit):
                continue  # can't classify this row's direction — skip rather than guess
            debit_paise = amount if is_debit else 0
            credit_paise = amount if is_credit else 0
        else:
            debit_paise = abs(_to_paise(col("debit")))
            credit_paise = abs(_to_paise(col("credit")))

        out.append(NormalizedTxn(
            transaction_date=iso,
            description=desc,
            reference_no=(str(ref).strip() or None) if ref is not None else None,
            debit_paise=debit_paise,
            credit_paise=credit_paise,
            balance_paise=_to_paise(col("balance")),
        ))
    return out


def _decode_csv(content: bytes) -> str:
    """Bytes to text the same way for inspection and for parsing.

    Extracted rather than repeated: if the mapper inspected a file as UTF-8 and
    the importer then read it as latin-1, the column the CA mapped and the
    column we read could differ on any row with a non-ASCII payee name.
    """
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("latin-1", errors="replace")


def _xlsx_rows(content: bytes) -> list[list]:
    """Non-empty rows from the active sheet. Same reason as _decode_csv."""
    try:
        from openpyxl import load_workbook  # lazy — only needed for XLSX
    except ImportError as e:  # pragma: no cover
        raise StatementParseError("XLSX support is unavailable on the server.") from e
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise StatementParseError("File is not a valid XLSX workbook.") from e
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return [r for r in rows if any(c is not None and str(c).strip() for c in r)]



# ── Explicit column mappings (audit Tier 3.2) ────────────────────────────────
#
# WHY THIS EXISTS
#     _ADAPTERS knows six layouts: HDFC, SBI, ICICI, Axis and two generic ones.
#     Every other bank — Kotak, IDFC First, PNB, Canara, Union, and every
#     co-operative bank in the country — hit `Unsupported bank statement
#     format` and stopped there. There was no way past it: the CA could not
#     tell us where the columns were, so a client banking anywhere unusual
#     could not get a statement into the product at all.
#
# WHY A MAPPING RATHER THAN EIGHT MORE ADAPTERS
#     An adapter is a guess about a layout nobody here has seen. Guessing eight
#     of them from memory is how you ship a Canara adapter that reads the
#     balance column as a credit — the F8 class of corruption _validate_adapter
#     was written about, arrived at deliberately. A mapping is not a guess: the
#     person who has the file in front of them says where the columns are, once,
#     and the answer is kept.
#
#     It also generalises. An adapter helps the banks somebody thought of; a
#     mapping helps every bank, including the ones nobody thought of.

#: The keys an adapter — and therefore a mapping — may carry.
MAPPING_KEYS = ("date", "desc", "ref", "debit", "credit", "amount", "drcr", "balance")
#: Without these two there is no transaction: no date, or nothing to call it.
REQUIRED_MAPPING_KEYS = ("date", "desc")


def validate_mapping(mapping: dict, header_count: int) -> dict[str, Optional[int]]:
    """A mapping the parser can use, or StatementParseError saying what is wrong.

    Structure only — deliberately NOT the column labels, since unrecognised
    labels are the whole reason a mapping is being supplied. The rules are the
    ones that make a mapping self-contradictory rather than merely unusual:

      * date and desc must be mapped — a row without them is not a transaction;
      * the amounts come EITHER as debit + credit columns OR as one amount
        column plus a Dr/Cr indicator, never as a mixture and never as neither.
        Half of each cannot be parsed by either branch of _rows_to_txns;
      * an index must exist in the file. A mapping pointing past the last
        column reads nothing, on every row, silently;
      * one column cannot be two things. Mapping `date` and `debit` to column 0
        is a slip, and the result — every amount unparseable, every row
        dropped — looks like an empty statement rather than a mistake.
    """
    if not isinstance(mapping, dict):
        raise StatementParseError("Column mapping must be an object of column positions.")

    clean: dict[str, Optional[int]] = {}
    for key, raw in mapping.items():
        if key not in MAPPING_KEYS:
            raise StatementParseError(
                f"Unknown column '{key}'. Expected one of: {', '.join(MAPPING_KEYS)}.")
        if raw is None:
            clean[key] = None
            continue
        try:
            idx = int(raw)
        except (TypeError, ValueError):
            raise StatementParseError(f"Column '{key}' must be a column position, got {raw!r}.")
        if idx < 0 or idx >= header_count:
            raise StatementParseError(
                f"Column '{key}' points at position {idx + 1}, but the file has "
                f"{header_count} column{'s' if header_count != 1 else ''}.")
        clean[key] = idx

    for key in REQUIRED_MAPPING_KEYS:
        if clean.get(key) is None:
            raise StatementParseError(
                f"The {key} column must be mapped — without it a row is not a transaction.")

    has_debit_credit = clean.get("debit") is not None or clean.get("credit") is not None
    has_amount_drcr = clean.get("amount") is not None or clean.get("drcr") is not None
    if has_debit_credit and has_amount_drcr:
        raise StatementParseError(
            "Map EITHER separate Debit and Credit columns, OR one Amount column "
            "with a Dr/Cr indicator — not both.")
    if not has_debit_credit and not has_amount_drcr:
        raise StatementParseError(
            "Map the amounts: either separate Debit and Credit columns, or one "
            "Amount column with a Dr/Cr indicator.")
    if has_amount_drcr and (clean.get("amount") is None or clean.get("drcr") is None):
        raise StatementParseError(
            "A single Amount column needs a Dr/Cr indicator column to say which "
            "way each row goes.")

    used = [(k, i) for k, i in clean.items() if i is not None]
    seen: dict[int, str] = {}
    for key, idx in used:
        if idx in seen:
            raise StatementParseError(
                f"Column {idx + 1} is mapped to both '{seen[idx]}' and '{key}'. "
                f"Each column can only be one thing.")
        seen[idx] = key

    # _rows_to_txns reads absent keys with .get(), so fill the shape out rather
    # than leaving it to every caller to remember which keys are optional.
    return {k: clean.get(k) for k in MAPPING_KEYS}


def header_fingerprint(headers: list) -> str:
    """A stable key for 'a file shaped like this one'.

    A saved mapping is reused only for a file whose header row still matches, so
    the fingerprint is what makes reuse SAFE rather than merely convenient. When
    a bank changes its export — a column inserted, one renamed — the
    fingerprint changes, the old mapping stops being applied, and the CA is
    asked once more. The alternative, keying on the account alone, silently
    reads the new layout with the old positions.

    Case- and whitespace-insensitive, because those vary between exports of the
    same layout and mean nothing.
    """
    import hashlib
    norm = "|".join(re.sub(r"\s+", " ", str(h or "")).strip().lower() for h in headers)
    return hashlib.sha256(norm.encode("utf-8")).hexdigest()[:32]


def balance_agreement(txns: list[NormalizedTxn]) -> dict:
    """Does the parse agree with the bank's own balance column?

    THE POINT OF THIS
        An explicit mapping skips the label check, so something else has to
        catch a mapping that parses cleanly and is wrong — debit and credit the
        wrong way round being the obvious one, and the most damaging: every
        amount is right, every date is right, and the direction of the client's
        entire cash position is inverted.

        The bank already told us the answer. An Indian statement carries a
        running balance, so between consecutive rows the balance MUST move by
        exactly the row's own movement. Swap debit and credit and every row
        disagrees at once.

    RELATION TO register.first_divergence
        Same idea, different arithmetic, and not a duplicate. That one checks a
        running balance built from a known opening figure, to find a missing or
        duplicated transaction in the BOOKS. This checks consecutive DELTAS in a
        file, needs no opening balance, and is asking whether we read the file
        correctly at all. Neither can be expressed as the other.

    A statement printed newest-first is a real and common export option, and it
    is not an error — every delta simply has the opposite sign. That is
    detected and reported rather than counted as disagreement, because telling
    a CA their mapping is wrong when it is right is how a correct mapping gets
    "fixed" into a broken one.
    """
    with_balance = [t for t in txns if t.balance_paise]
    if len(with_balance) < 2:
        return {"checked": False, "reason": "the statement has no balance column to check against"}

    def _disagreements(seq: list[NormalizedTxn]) -> list[dict]:
        out = []
        for prev, cur in zip(seq, seq[1:]):
            expected = cur.credit_paise - cur.debit_paise
            actual = cur.balance_paise - prev.balance_paise
            if expected != actual:
                out.append({
                    "transaction_date": cur.transaction_date,
                    "description": cur.description[:80],
                    "expected_movement_paise": expected,
                    "balance_moved_paise": actual,
                })
        return out

    forward = _disagreements(with_balance)
    if not forward:
        return {"checked": True, "agrees": True, "rows_checked": len(with_balance) - 1,
                "order": "oldest-first"}

    reverse = _disagreements(list(reversed(with_balance)))
    if not reverse:
        return {"checked": True, "agrees": True, "rows_checked": len(with_balance) - 1,
                "order": "newest-first",
                "note": "This statement is printed newest-first. The balances agree."}

    checked = len(with_balance) - 1
    worse = forward if len(forward) <= len(reverse) else reverse
    return {
        "checked": True,
        "agrees": False,
        "rows_checked": checked,
        "disagreeing_rows": len(worse),
        "first_disagreement": worse[0],
        "reason": (
            f"{len(worse)} of {checked} rows do not move the bank's balance by their "
            f"own amount. The most likely cause is a column mapped to the wrong "
            f"thing — check Debit and Credit are not swapped."),
    }


def inspect_statement(filename: str, content: bytes, *, sample_rows: int = 8) -> dict:
    """What the mapping screen needs: the header row, some real rows, and a start.

    Deliberately does NOT parse into transactions. The file is here precisely
    because parsing it failed, so anything that needs a working mapping has to
    come after the CA supplies one.

    `detected_format` is what detect_format WOULD choose; `detected_fits` says
    whether that choice survives _validate_adapter. The pair matters: a file can
    be detected as 'hdfc' on a shared cheque-column signal and still not be an
    HDFC statement, which is the case this whole feature exists to serve.
    """
    name = (filename or "").lower().strip()
    if name.endswith(".csv"):
        rows = [r for r in csv.reader(io.StringIO(_decode_csv(content)))
                if any(str(c).strip() for c in r)]
    elif name.endswith(".xlsx"):
        rows = _xlsx_rows(content)
    else:
        raise StatementParseError(
            "Unsupported file type — upload a .csv or .xlsx bank statement.")
    if not rows:
        raise StatementParseError("File has no data rows.")

    header_idx = _find_header_idx(rows)
    headers = [str(c) if c is not None else "" for c in rows[header_idx]]
    body = rows[header_idx + 1:]

    fmt = detect_format(headers)
    try:
        _validate_adapter(fmt, headers)
        fits = True
    except StatementParseError:
        fits = False

    return {
        "headers": headers,
        "sample_rows": [[("" if c is None else str(c)) for c in r][:len(headers)]
                        for r in body[:sample_rows]],
        "total_rows": len(body),
        "header_row_index": header_idx,
        "detected_format": fmt,
        "detected_fits": fits,
        # Only offer the detected layout as a starting point when it actually
        # fits. Prefilling a mapping we have just established is wrong would be
        # handing the CA the error to confirm.
        "proposed_mapping": {k: _ADAPTERS[fmt].get(k) for k in MAPPING_KEYS} if fits else None,
        "header_fingerprint": header_fingerprint(headers),
    }


def _find_header_idx(rows: list[list]) -> int:
    for i, cells in enumerate(rows[:10]):
        if _looks_like_header([str(c) for c in cells]):
            return i
    return 0


def parse_csv(text: str, mapping: Optional[dict] = None) -> list[NormalizedTxn]:
    text = text.lstrip("﻿")  # strip BOM
    reader = list(csv.reader(io.StringIO(text)))
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        raise StatementParseError("File has no data rows.")
    txns = _rows_to_txns(rows, _find_header_idx(rows), mapping)
    if not txns:
        raise StatementParseError("No transactions found — check the file format/columns.")
    return txns


def parse_xlsx(content: bytes, mapping: Optional[dict] = None) -> list[NormalizedTxn]:
    rows = _xlsx_rows(content)
    if len(rows) < 2:
        raise StatementParseError("Workbook has no data rows.")
    txns = _rows_to_txns(rows, _find_header_idx(rows), mapping)
    if not txns:
        raise StatementParseError("No transactions found — check the file format/columns.")
    return txns


def parse_statement(filename: str, content: bytes,
                    mapping: Optional[dict] = None) -> list[NormalizedTxn]:
    """Dispatch by extension. Raises StatementParseError on unsupported/malformed.

    `mapping` is an explicit column mapping (Tier 3.2), used INSTEAD of
    detect_format when the CA has told us where this bank's columns are. None
    keeps the original behaviour exactly.
    """
    name = (filename or "").lower().strip()
    if name.endswith(".csv"):
        return parse_csv(_decode_csv(content), mapping)
    if name.endswith(".xlsx"):
        return parse_xlsx(content, mapping)
    raise StatementParseError("Unsupported file type — upload a .csv or .xlsx bank statement.")
